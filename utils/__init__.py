# -*- coding: utf-8 -*-
"""
工具函数：文件内容提取、LLM调用等
"""
import os
import json
import requests

# 支持的文件格式
ALLOWED_EXTENSIONS = {'txt', 'md', 'pdf', 'docx', 'xlsx', 'xls', 'doc', 'html', 'htm', 'chm', 'log', 'csv'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def safe_filename(filename):
    """安全化上传文件名：保留中文，去除路径分隔与危险字符

    仅影响新上传文件的落盘名称，不保证与原始名一致。
    """
    if not filename:
        return 'unnamed_file'
    # 去掉路径部分（兼容 / 与 \）
    filename = filename.replace('\\', '/').rsplit('/', 1)[-1]
    if filename in ('.', '..'):
        return 'unnamed_file'
    import re
    # 仅保留字母/数字/中文/下划线/连字符/点/空格
    filename = re.sub(r'[^\w一-鿿\-\. ]', '_', filename)
    # 去掉开头的点，防止隐藏文件与相对路径
    filename = filename.lstrip('.')
    # 折叠连续下划线并清理首尾
    filename = re.sub(r'_+', '_', filename).strip('_').strip()
    return filename if filename else 'unnamed_file'


def safe_join(base_dir, *parts):
    """拼接路径并校验结果仍位于 base_dir 内，越界返回 None

    Args:
        base_dir: 允许的根目录（应为绝对路径）
        parts: 待拼接的子路径段，可包含用户输入

    Returns:
        str: 校验通过后的绝对路径；若解析后越出 base_dir 返回 None
    """
    filepath = os.path.realpath(os.path.join(base_dir, *parts))
    rel = os.path.relpath(filepath, base_dir)
    if rel.startswith('..' + os.sep) or rel == '..' or os.path.isabs(rel):
        return None
    return filepath


_CRED_KEY_CACHE = None


def _credential_key():
    """由 SECRET_KEY 派生 Fernet 密钥（PBKDF2-SHA256），结果缓存"""
    global _CRED_KEY_CACHE
    if _CRED_KEY_CACHE:
        return _CRED_KEY_CACHE
    import hashlib
    import base64
    from config import SECRET_KEY
    derived = hashlib.pbkdf2_hmac(
        'sha256', SECRET_KEY.encode('utf-8'), b'dbsv_cred', 100000)
    _CRED_KEY_CACHE = base64.urlsafe_b64encode(derived)
    return _CRED_KEY_CACHE


def encrypt_secret(plaintext):
    """加密凭据（Fernet），cryptography 不可用或加密失败时记警告并回退明文

    Args:
        plaintext: 待加密的明文（密码/私钥/passphrase）

    Returns:
        str: 密文 token；回退时返回原明文
    """
    if not plaintext:
        return plaintext
    try:
        from cryptography.fernet import Fernet
    except ImportError:
        print("[utils] 警告: cryptography 未安装，凭据将以明文存储")
        return plaintext
    try:
        f = Fernet(_credential_key())
        return f.encrypt(plaintext.encode('utf-8')).decode('utf-8')
    except Exception as e:
        print(f"[utils] 凭据加密失败，回退明文: {e}")
        return plaintext


def decrypt_secret(token):
    """解密凭据；无法解密（旧明文或密钥变更）时原样返回

    Args:
        token: Fernet 密文 token 或历史明文

    Returns:
        str: 解密后的明文；无法解密时返回原值
    """
    if not token:
        return token
    try:
        from cryptography.fernet import Fernet
    except ImportError:
        return token
    try:
        f = Fernet(_credential_key())
        return f.decrypt(token.encode('utf-8')).decode('utf-8')
    except Exception:
        return token


def extract_content(filepath):
    """根据文件扩展名提取文本内容"""
    ext = filepath.rsplit('.', 1)[-1].lower() if '.' in filepath else ''

    extractors = {
        'txt': _extract_txt,
        'md': _extract_txt,
        'html': _extract_txt,
        'htm': _extract_txt,
        'log': _extract_txt,
        'csv': _extract_txt,
        'pdf': _extract_pdf,
        'docx': _extract_docx,
        'doc': _extract_docx,
        'xlsx': _extract_xlsx,
        'xls': _extract_xlsx,
        'chm': _extract_chm,
    }

    extractor = extractors.get(ext)
    if extractor:
        try:
            return extractor(filepath)
        except Exception as e:
            return f'[解析失败: {str(e)}]'
    return ''


def _extract_txt(filepath):
    encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']
    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                return f.read()
        except (UnicodeDecodeError, UnicodeError):
            continue
    return ''


def _extract_pdf(filepath):
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(filepath)
        text_parts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
        return '\n'.join(text_parts)
    except ImportError:
        return '[PyPDF2 未安装，无法解析 PDF]'
    except Exception as e:
        return f'[PDF 解析失败: {str(e)}]'


def _extract_docx(filepath):
    try:
        from docx import Document
        doc = Document(filepath)
        return '\n'.join(para.text for para in doc.paragraphs if para.text.strip())
    except ImportError:
        return '[python-docx 未安装，无法解析 DOCX]'
    except Exception as e:
        return f'[DOCX 解析失败: {str(e)}]'


def _extract_xlsx(filepath):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        text_parts = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = ' '.join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text_parts.append(row_text)
        wb.close()
        return '\n'.join(text_parts)
    except ImportError:
        return '[openpyxl 未安装，无法解析 XLSX]'
    except Exception as e:
        return f'[XLSX 解析失败: {str(e)}]'


def _extract_chm(filepath):
    """提取CHM文件中的文本内容"""
    try:
        import subprocess
        import tempfile
        import os

        # 创建临时目录用于解压CHM
        with tempfile.TemporaryDirectory() as tmpdir:
            # 尝试使用7z解压CHM文件
            try:
                result = subprocess.run(
                    ['7z', 'x', filepath, f'-o{tmpdir}', '-y'],
                    capture_output=True,
                    text=True,
                    timeout=30
                )
                if result.returncode != 0:
                    return f'[CHM 解压失败: {result.stderr[:200]}]'
            except FileNotFoundError:
                return '[CHM 解析需要 7z 工具，请安装 7-Zip]'
            except Exception as e:
                return f'[CHM 解压失败: {str(e)}]'

            # 递归读取所有HTML文件内容
            text_parts = []
            for root, dirs, files in os.walk(tmpdir):
                for filename in files:
                    if filename.lower().endswith(('.html', '.htm')):
                        file_path = os.path.join(root, filename)
                        try:
                            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                                content = f.read()
                                # 简单去除HTML标签
                                import re
                                text = re.sub(r'<[^>]+>', ' ', content)
                                text = re.sub(r'\s+', ' ', text).strip()
                                if text:
                                    text_parts.append(text)
                        except Exception:
                            pass

            if text_parts:
                return '\n\n'.join(text_parts)
            return '[CHM 文件中未找到可提取的文本内容]'
    except Exception as e:
        return f'[CHM 解析失败: {str(e)}]'


def load_llm_config(model_id=None):
    """加载大模型API配置

    Args:
        model_id: 指定模型ID，如果不指定则使用默认模型
    """
    from db.database import get_config
    import json

    models = get_config('llm_models', [])
    if not isinstance(models, list):
        models = []

    target_model = None

    if model_id and models:
        # 查找指定模型
        for model in models:
            if model.get('id') == model_id:
                target_model = model
                break

    if not target_model and models:
        # 查找默认模型
        default_id = get_config('default_model_id', '')
        for model in models:
            if model.get('id') == default_id:
                target_model = model
                break
        # 如果没有默认模型，使用第一个
        if not target_model:
            target_model = models[0]

    if target_model:
        return {
            'api_url': target_model.get('api_url', ''),
            'api_key': target_model.get('api_key', ''),
            'model_name': target_model.get('model_name', ''),
            'model_id': target_model.get('id', '')
        }

    # 兼容旧配置
    return {
        'api_url': get_config('api_url', ''),
        'api_key': get_config('api_key', ''),
        'model_name': get_config('model_name', ''),
        'model_id': ''
    }


def _build_api_url(config):
    """根据配置构建完整的 API URL"""
    api_url = config.get('api_url', '').rstrip('/')
    if api_url.endswith('/v1'):
        api_url += '/chat/completions'
    elif not api_url.endswith('/chat/completions'):
        api_url += '/chat/completions'
    return api_url


def _build_api_headers(config):
    """构建 API 请求头"""
    return {
        "Authorization": f"Bearer {config.get('api_key', '')}",
        "Content-Type": "application/json"
    }


def _build_api_data(config, messages, stream=False):
    """构建 API 请求数据"""
    model = config.get('model_name', 'gpt-3.5-turbo')
    # Moonshot kimi-k2.6 模型只支持 temperature=1
    if 'kimi' in model.lower():
        temperature = 1
    else:
        temperature = 0.7
    return {
        "model": model,
        "messages": messages,
        "temperature": temperature,
        "stream": stream
    }


def _check_llm_config(config):
    """检查 LLM 配置是否完整"""
    if not config.get('api_url') or not config.get('api_key'):
        return False, "请先配置大模型API信息"
    return True, None


def call_llm(messages, model_id=None):
    """调用大模型API（兼容OpenAI格式）

    Args:
        messages: 消息列表
        model_id: 指定模型ID，如果不指定则使用默认模型
    """
    config = load_llm_config(model_id)

    ok, error = _check_llm_config(config)
    if not ok:
        return None, error

    api_url = _build_api_url(config)
    headers = _build_api_headers(config)
    data = _build_api_data(config, messages)

    try:
        response = requests.post(
            api_url,
            json=data,
            headers=headers,
            timeout=120
        )
        response.raise_for_status()
        result = response.json()
        return result['choices'][0]['message']['content'], None
    except requests.exceptions.HTTPError as e:
        # 尝试获取响应体中的详细错误信息
        error_detail = str(e)
        try:
            error_json = response.json()
            error_detail = f"{error_detail} | 响应: {json.dumps(error_json, ensure_ascii=False)}"
        except:
            try:
                error_text = response.text[:500]
                error_detail = f"{error_detail} | 响应: {error_text}"
            except:
                pass
        return None, f"API调用失败: {error_detail}"
    except requests.exceptions.RequestException as e:
        return None, f"API调用失败: {str(e)}"
    except (KeyError, IndexError) as e:
        return None, f"API响应格式错误: {str(e)}"


def call_llm_stream(messages, model_id=None):
    """调用大模型API（流式输出，兼容OpenAI格式）

    Args:
        messages: 消息列表
        model_id: 指定模型ID，如果不指定则使用默认模型
    """
    config = load_llm_config(model_id)

    ok, error = _check_llm_config(config)
    if not ok:
        yield None, error
        return

    api_url = _build_api_url(config)
    headers = _build_api_headers(config)
    data = _build_api_data(config, messages, stream=True)

    try:
        response = requests.post(
            api_url,
            json=data,
            headers=headers,
            timeout=120,
            stream=True
        )
        response.raise_for_status()

        for line in response.iter_lines():
            if line:
                line = line.decode('utf-8')
                if line.startswith('data: '):
                    line = line[6:]
                    if line.strip() == '[DONE]':
                        break
                    try:
                        chunk = json.loads(line)
                        if 'choices' in chunk and len(chunk['choices']) > 0:
                            delta = chunk['choices'][0].get('delta', {})
                            if 'content' in delta:
                                yield delta['content'], None
                    except json.JSONDecodeError:
                        continue
    except requests.exceptions.RequestException as e:
        yield None, f"API调用失败: {str(e)}"


def stream_llm_response(messages, model_id=None):
    """通用 LLM 流式响应生成器，返回 SSE 格式数据

    Args:
        messages: 消息列表
        model_id: 指定模型ID，如果不指定则使用默认模型

    Yields:
        str: SSE 格式数据行
    """
    for content, error in call_llm_stream(messages, model_id=model_id):
        if error:
            yield f"data: {json.dumps({'error': error})}\n\n"
            break
        if content:
            yield f"data: {json.dumps({'content': content})}\n\n"
    yield "data: [DONE]\n\n"
