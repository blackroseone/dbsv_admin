# -*- coding: utf-8 -*-
"""
集群拓扑批量导入模块
支持从Excel文件导入服务器清单和实例清单
"""
import uuid
import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _is_example_row(ws, row_idx, headers):
    """
    判断某一行是否是示例数据行
    示例数据的特征：
    1. 字体是斜体（italic=True）
    2. 或者单元格背景色是灰色（F0F0F0）
    """
    # 检查第一个有值的单元格的字体和填充色
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:  # 有值的单元格
            # 检查字体是否为斜体
            if cell.font and cell.font.italic:
                return True
            # 检查背景色是否为灰色（示例数据的标记）
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == '00F0F0F0':
                return True
    return False


def import_servers_from_excel(filepath, db_module):
    """
    从Excel导入服务器清单
    返回: (success_count, error_list)
    """
    conn = db_module.get_db()
    success_count = 0
    errors = []

    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb['服务器清单']

        # 读取表头（第3行是字段名）
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=3, column=col).value
            headers.append(val.strip() if val else '')

        # 数据从第5行开始（第4行是说明，第5行开始是数据）
        for row_idx in range(5, ws.max_row + 1):
            # 跳过示例数据行
            if _is_example_row(ws, row_idx, headers):
                continue

            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = str(val).strip() if val is not None else ''

            # 跳过空行
            if not row_data.get('pool_name'):
                continue

            try:
                _import_single_server(row_data, conn)
                success_count += 1
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")
                logger.warning(f"导入服务器失败 第{row_idx}行: {e}")

        conn.commit()
    except Exception as e:
        errors.append(f"解析Excel失败: {str(e)}")
        logger.error(f"解析服务器清单Excel失败: {e}")

    return success_count, errors


def _import_single_server(data, conn):
    """导入单个服务器（包含资源池、集群）"""
    pool_name = data.get('pool_name', '').strip()
    pool_db_type = data.get('pool_db_type', 'MySQL').strip()
    pool_env = data.get('pool_env', 'production').strip()
    cluster_name = data.get('cluster_name', '').strip()
    server_name = data.get('server_name', '').strip()
    sn = data.get('sn', '').strip()
    ip = data.get('ip', '').strip()
    datacenter = data.get('datacenter', '').strip()
    node_role = data.get('node_role', '计算节点').strip()
    hardware_type = data.get('hardware_type', '非信创物理机').strip()
    cpu = data.get('cpu', '').strip()
    memory = data.get('memory', '').strip()
    description = data.get('description', '').strip()

    if not pool_name:
        raise ValueError("资源池名称不能为空")
    if not server_name:
        raise ValueError("服务器名称不能为空")
    if not ip:
        raise ValueError("IP地址不能为空")

    # 1. 查找或创建资源池
    pool_row = conn.execute(
        "SELECT id FROM resource_pools WHERE name=?",
        (pool_name,)
    ).fetchone()

    if pool_row:
        pool_id = pool_row['id']
    else:
        pool_id = str(uuid.uuid4())
        conn.execute(
            "INSERT INTO resource_pools (id, name, db_type, environment, description) VALUES (?, ?, ?, ?, ?)",
            (pool_id, pool_name, pool_db_type, pool_env, '')
        )

    # 2. 查找或创建集群（如果有指定）
    cluster_id = ''
    if cluster_name:
        cluster_row = conn.execute(
            "SELECT id FROM clusters WHERE name=? AND resource_pool_id=?",
            (cluster_name, pool_id)
        ).fetchone()

        if cluster_row:
            cluster_id = cluster_row['id']
        else:
            cluster_id = str(uuid.uuid4())
            conn.execute(
                "INSERT INTO clusters (id, resource_pool_id, name, description) VALUES (?, ?, ?, ?)",
                (cluster_id, pool_id, cluster_name, '')
            )

    # 3. 查找或创建服务器（按IP去重）
    server_row = conn.execute(
        "SELECT id FROM servers WHERE host=?",
        (ip,)
    ).fetchone()

    if server_row:
        # 更新现有服务器
        conn.execute(
            """UPDATE servers SET
                resource_pool_id=?, cluster_id=?, name=?, sn=?, datacenter=?,
                node_role=?, hardware_type=?, cpu=?, memory=?, description=?
            WHERE host=?""",
            (pool_id, cluster_id, server_name, sn, datacenter,
             node_role, hardware_type, cpu, memory, description, ip)
        )
    else:
        # 创建新服务器
        server_id = str(uuid.uuid4())
        conn.execute(
            """INSERT INTO servers
                (id, resource_pool_id, cluster_id, name, sn, host, datacenter,
                 node_role, hardware_type, cpu, memory, description)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (server_id, pool_id, cluster_id, server_name, sn, ip, datacenter,
             node_role, hardware_type, cpu, memory, description)
        )


def import_instances_from_excel(filepath, db_module):
    """
    从Excel导入实例清单
    返回: (success_count, error_list)
    """
    conn = db_module.get_db()
    success_count = 0
    errors = []

    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb['实例清单']

        # 读取表头（第3行是字段名）
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=3, column=col).value
            headers.append(val.strip() if val else '')

        # 数据从第5行开始
        for row_idx in range(5, ws.max_row + 1):
            # 跳过示例数据行
            if _is_example_row(ws, row_idx, headers):
                continue

            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = str(val).strip() if val is not None else ''

            # 跳过空行
            if not row_data.get('ip'):
                continue

            try:
                _import_single_instance(row_data, conn)
                success_count += 1
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")
                logger.warning(f"导入实例失败 第{row_idx}行: {e}")

        conn.commit()
    except Exception as e:
        errors.append(f"解析Excel失败: {str(e)}")
        logger.error(f"解析实例清单Excel失败: {e}")

    return success_count, errors


def _import_single_instance(data, conn):
    """导入单个实例（包含租户）"""
    ip = data.get('ip', '').strip()
    tenant_name = data.get('tenant_name', '').strip()
    tenant_topology = data.get('tenant_topology', 'master-slave').strip()
    tenant_spec = data.get('tenant_spec', 'small-8c32g').strip()
    instance_name = data.get('instance_name', '').strip()
    port = data.get('port', '3306').strip()
    role = data.get('role', 'slave').strip()
    cpu = data.get('cpu', '').strip()
    memory = data.get('memory', '').strip()
    description = data.get('description', '').strip()

    if not ip:
        raise ValueError("IP地址不能为空")
    if not tenant_name:
        raise ValueError("租户名称不能为空")
    if not instance_name:
        raise ValueError("实例名称不能为空")

    # 1. 通过IP查找服务器
    server_row = conn.execute(
        "SELECT id, resource_pool_id FROM servers WHERE host=?",
        (ip,)
    ).fetchone()

    if not server_row:
        raise ValueError(f"找不到IP为 {ip} 的服务器，请先导入服务器清单")

    server_id = server_row['id']
    resource_pool_id = server_row['resource_pool_id']

    # 2. 查找或创建租户（按名称+资源池去重）
    tenant_row = conn.execute(
        "SELECT id FROM tenants WHERE name=? AND resource_pool_id=?",
        (tenant_name, resource_pool_id)
    ).fetchone()

    if tenant_row:
        tenant_id = tenant_row['id']
    else:
        tenant_id = str(uuid.uuid4())
        conn.execute(
            "INSERT INTO tenants (id, resource_pool_id, name, topology_type, spec, description) VALUES (?, ?, ?, ?, ?, ?)",
            (tenant_id, resource_pool_id, tenant_name, tenant_topology, tenant_spec, '')
        )

    # 3. 查找或创建实例（按服务器+名称+端口去重）
    instance_row = conn.execute(
        "SELECT id FROM instances WHERE server_id=? AND name=? AND port=?",
        (server_id, instance_name, port)
    ).fetchone()

    if instance_row:
        # 更新现有实例
        conn.execute(
            """UPDATE instances SET
                tenant_id=?, role=?, cpu=?, memory=?, description=?
            WHERE id=?""",
            (tenant_id, role, cpu, memory, description, instance_row['id'])
        )
    else:
        # 创建新实例
        instance_id = str(uuid.uuid4())
        conn.execute(
            """INSERT INTO instances
                (id, server_id, tenant_id, name, port, cpu, memory, role, tenant_role, description)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (instance_id, server_id, tenant_id, instance_name, port,
             cpu, memory, role, role, description)
        )
