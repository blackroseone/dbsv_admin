# -*- coding: utf-8 -*-
"""
生成集群拓扑数据批量导入模板 (Excel) - 简化版
两个sheet页：服务器清单 + 实例清单
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def create_import_template(output_path):
    """创建集群拓扑批量导入模板"""
    wb = Workbook()

    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    required_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ============ 工作表1: 服务器清单 ============
    ws_server = wb.active
    ws_server.title = "服务器清单"
    server_headers = [
        ("pool_name", "资源池名称", "必填，如: 生产环境"),
        ("pool_db_type", "数据库类型", "如: MySQL, Oracle, PostgreSQL, MongoDB"),
        ("pool_env", "环境", "如: production, development, test"),
        ("cluster_name", "集群名称", "可选，如: 主库集群A"),
        ("server_name", "服务器名称", "可选，留空则根据IP自动生成，如: node-192.168.1.10"),
        ("sn", "SN序列号", "可选，服务器硬件序列号"),
        ("ip", "IP地址", "必填，唯一标识服务器，如: 192.168.1.10"),
        ("datacenter", "数据中心", "可选，如: 北京机房"),
        ("node_role", "节点角色", "如: 计算节点, 存储节点, 管理节点"),
        ("hardware_type", "硬件类型", "如: 非信创物理机, 信创物理机, 虚拟机"),
        ("cpu", "CPU", "可选，如: 32核"),
        ("memory", "内存", "可选，如: 128GB"),
        ("description", "描述", "可选"),
    ]
    _write_sheet(ws_server, server_headers, "服务器清单", header_fill, header_font, required_fill, thin_border)
    _add_example(ws_server, [
        "示例数据-生产环境", "MySQL", "production",
        "主库集群A", "db-server-01", "SN123456789", "192.168.1.10",
        "北京机房", "计算节点", "非信创物理机", "32核", "128GB", "主库服务器"
    ])
    _add_example(ws_server, [
        "示例数据-生产环境", "MySQL", "production",
        "从库集群B", "db-server-02", "SN987654321", "192.168.1.11",
        "北京机房", "计算节点", "非信创物理机", "32核", "128GB", "从库服务器"
    ])
    _add_example(ws_server, [
        "示例数据-测试环境", "MySQL", "test",
        "", "test-server-01", "", "192.168.2.10",
        "上海机房", "计算节点", "虚拟机", "16核", "64GB", "测试服务器"
    ])

    # 添加下拉框数据验证
    _add_dropdown_validation(ws_server, "B", "Oracle,PostgreSQL,MongoDB,GoldenDB,OceanBase,GaussDB,DM,TDSQL,MySQL", 5)
    _add_dropdown_validation(ws_server, "C", "production,dev-test,uat-prod", 5)
    _add_dropdown_validation(ws_server, "I", "计算节点,存储节点,管理节点", 5)
    _add_dropdown_validation(ws_server, "J", "非信创物理机,非信创虚拟机,海光物理机,海光虚拟机,鲲鹏物理机,鲲鹏虚拟机", 5)

    # ============ 工作表2: 实例清单 ============
    ws_instance = wb.create_sheet("实例清单")
    instance_headers = [
        ("ip", "IP地址", "必填，关联到服务器清单中的IP"),
        ("tenant_name", "租户名称", "必填，如: 业务系统A"),
        ("tenant_topology", "租户拓扑类型", "如: master-slave, single, mha, mgr"),
        ("tenant_spec", "租户规格", "如: small-8c32g, medium-16c64g, large-32c128g"),
        ("instance_name", "实例名称", "必填，如: mysql-master-01"),
        ("port", "端口", "如: 3306, 1521, 5432, 27017"),
        ("role", "实例角色", "如: master, slave, standalone"),
        ("cpu", "CPU", "可选，如: 8核"),
        ("memory", "内存", "可选，如: 32GB"),
        ("description", "描述", "可选"),
    ]
    _write_sheet(ws_instance, instance_headers, "实例清单", header_fill, header_font, required_fill, thin_border)
    _add_example(ws_instance, [
        "示例数据-192.168.1.10", "业务系统A", "master-slave", "medium-16c64g",
        "mysql-master-01", "3306", "master", "8核", "32GB", "主实例"
    ])
    _add_example(ws_instance, [
        "示例数据-192.168.1.11", "业务系统A", "master-slave", "medium-16c64g",
        "mysql-slave-01", "3306", "slave", "8核", "32GB", "从实例"
    ])
    _add_example(ws_instance, [
        "示例数据-192.168.2.10", "测试业务", "single", "small-8c32g",
        "mysql-single-01", "3306", "standalone", "4核", "16GB", "单实例"
    ])

    # 添加下拉框数据验证
    _add_dropdown_validation(ws_instance, "C", "master-slave,single,mha,paxos/raft,rac", 5)
    _add_dropdown_validation(ws_instance, "D", "macro-1c4g,macro-2c8g,macro-4c16g,small-8c16g,small-16c64g,medium-32C128G,large-64c256g,exlu-128c512g", 5)
    _add_dropdown_validation(ws_instance, "G", "master,slave,standalone", 5)

    # ============ 工作表3: 导入说明 ============
    ws_guide = wb.create_sheet("导入说明")
    _write_guide(ws_guide, header_fill, header_font, thin_border)

    # 保存
    wb.save(output_path)
    print(f"[OK] Import template generated: {output_path}")
    return output_path


def _add_dropdown_validation(ws, col_letter, options, start_row):
    """添加下拉框数据验证"""
    dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
    dv.error = '请从下拉列表中选择'
    dv.errorTitle = '无效输入'
    dv.prompt = '请从列表中选择'
    dv.promptTitle = '下拉选项'
    ws.add_data_validation(dv)
    # 应用到从start_row开始的1000行
    dv.add(f'{col_letter}{start_row}:{col_letter}1000')


def _write_sheet(ws, headers, sheet_name, header_fill, header_font, required_fill, thin_border):
    """写入表头和说明"""
    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"{sheet_name} - 批量导入数据")
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 表头
    for col_idx, (field, name, desc) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # 字段名行（隐藏，用于程序识别）
    for col_idx, (field, name, desc) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=field)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(italic=True, size=9, color="666666")
        cell.border = thin_border

    # 说明行
    for col_idx, (field, name, desc) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=desc)
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        cell.font = Font(size=9, color="333333")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 40


def _add_example(ws, values):
    """添加示例数据行（标记为灰色斜体）"""
    row = ws.max_row + 1
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.font = Font(color="999999", italic=True)
        # 添加标记表示这是示例数据
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")


def _write_guide(ws, header_fill, header_font, thin_border):
    """写入导入说明"""
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 85

    guide_data = [
        ("导入步骤", ""),
        ("", ""),
        ("1. 填写服务器清单", "在'服务器清单'工作表中填写所有服务器信息。IP地址必填且唯一，用于关联实例。"),
        ("2. 填写实例清单", "在'实例清单'工作表中填写所有实例信息。通过IP地址关联到服务器。"),
        ("3. 执行导入", "通过系统批量导入功能上传此Excel文件，系统自动处理关联关系。"),
        ("", ""),
        ("服务器清单说明", ""),
        ("", ""),
        ("资源池名称", "必填。相同名称的资源池会被合并。系统会自动创建资源池并生成ID。"),
        ("数据库类型", "如: MySQL, Oracle, PostgreSQL, MongoDB。用于标识资源池的数据库类型。"),
        ("环境", "如: production, development, test。标识资源池的环境类型。"),
        ("集群名称", "可选。相同名称+相同资源池的集群会被合并。留空表示服务器不归属任何集群。"),
        ("服务器名称", "必填。建议有意义的命名，如 db-server-01。"),
        ("IP地址", "必填且唯一。作为服务器的主键，也用于实例关联。格式如 192.168.1.10。"),
        ("", ""),
        ("实例清单说明", ""),
        ("", ""),
        ("IP地址", "必填。必须对应服务器清单中已存在的IP地址。系统通过此字段查找关联的服务器。"),
        ("租户名称", "必填。相同名称+相同资源池的租户会被合并。系统会自动创建租户并生成ID。"),
        ("租户拓扑类型", "如: master-slave, single, mha, mgr。标识租户的架构类型。"),
        ("租户规格", "如: small-8c32g, medium-16c64g, large-32c128g。标识租户的资源规格。"),
        ("实例名称", "必填。建议有意义的命名，如 mysql-master-01。"),
        ("端口", "数据库监听端口。如: 3306(MySQL), 1521(Oracle), 5432(PostgreSQL)。"),
        ("实例角色", "如: master, slave, standalone。标识实例在集群中的角色。"),
        ("", ""),
        ("导入规则", ""),
        ("", ""),
        ("ID自动生成", "所有ID（资源池、集群、服务器、租户、实例）均由系统自动生成，无需手动填写。"),
        ("重复处理", "资源池/集群/租户按名称去重（同一资源池内）。服务器按IP去重。"),
        ("关联自动填充", "实例会自动关联到对应的服务器、租户和资源池，无需手动指定ID。"),
        ("导入顺序", "先导入服务器清单，再导入实例清单。系统会自动处理依赖关系。"),
        ("", ""),
        ("示例数据", "工作表中的灰色斜体数据为示例，请替换为您的实际业务数据。"),
    ]

    for row_idx, (key, val) in enumerate(guide_data, 1):
        cell_a = ws.cell(row=row_idx, column=1, value=key)
        cell_b = ws.cell(row=row_idx, column=2, value=val)
        for cell in (cell_a, cell_b):
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if key in ("导入步骤", "服务器清单说明", "实例清单说明", "导入规则"):
            cell_a.fill = header_fill
            cell_a.font = header_font
            ws.row_dimensions[row_idx].height = 25
        elif key:
            cell_a.font = Font(bold=True)
            ws.row_dimensions[row_idx].height = 22


if __name__ == "__main__":
    output = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cluster_topology_import_template_v2.xlsx")
    create_import_template(output)
